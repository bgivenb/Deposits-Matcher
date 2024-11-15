import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import Font
from itertools import combinations
from PIL import Image, ImageTk
from itertools import product
import threading

import sys
import os


# Program: DepositsMatcher
# Author: Given Borthwick

class ScrollableFrame(tk.Frame):
    """
    A scrollable frame that can be used to contain multiple widgets.
    """
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        # Create a canvas and a vertical scrollbar for scrolling it
        canvas = tk.Canvas(self, borderwidth=0, background="#303030")  # Dark background
        scrollbar = tk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = tk.Frame(canvas, background="#303030")  # Dark background

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

class DepositsMatcherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("DepositsMatcher - by Given Borthwick")
        self.root.geometry("900x600")  # Increased window size for better visibility
        self.root.configure(bg="#303030")  # Set main window background to dark gray

        # Title with icon at the top
        title_frame = tk.Frame(root, bg="#303030")
        title_frame.pack(padx=10, pady=10, fill="x")

        # Text "Deposits" Label
        title_label_left = tk.Label(title_frame, text="Deposits", font=("Arial", 16, "bold"), bg="#303030", fg="white")
        title_label_left.pack(side=tk.LEFT)

        # Determine the path for the icon image
        if hasattr(sys, '_MEIPASS'):
            icon_path = os.path.join(sys._MEIPASS, "deposits.png")
        else:
            icon_path = "deposits.png"

        # Load and display icon in the title
        try:
            icon_image = Image.open(icon_path)  # Load the image
            icon_image = icon_image.resize((24, 24), Image.LANCZOS)  # Resize with high-quality resampling
            icon_photo = ImageTk.PhotoImage(icon_image)  # Convert to PhotoImage for Tkinter
            icon_label = tk.Label(title_frame, image=icon_photo, bg="#303030")
            icon_label.image = icon_photo  # Keep a reference to avoid garbage collection
            icon_label.pack(side=tk.LEFT, padx=5)
        except Exception as e:
            print("Error loading icon:", e)

        # Text "Matcher" Label
        title_label_right = tk.Label(title_frame, text="Matcher", font=("Arial", 16, "bold"), bg="#303030", fg="white")
        title_label_right.pack(side=tk.LEFT)

        # Help button
        help_button = tk.Button(title_frame, text="Help", command=self.show_help)  # Default styling
        help_button.pack(side=tk.RIGHT, padx=10)

        # Input frame for number of deposits
        input_frame = tk.Frame(root, bg="#303030")
        input_frame.pack(padx=10, pady=5, fill="x")

        # Input for number of deposits for List A
        tk.Label(input_frame, text="Number of Deposits for List A:", bg="#303030", fg="white").grid(row=0, column=0, padx=10, pady=5, sticky="e")
        self.num_deposits_a_entry = tk.Entry(input_frame, bg="#1d1d1e", fg="white")
        self.num_deposits_a_entry.grid(row=0, column=1, padx=10, pady=5, sticky="w")

        # Input for number of deposits for List B
        tk.Label(input_frame, text="Number of Deposits for List B:", bg="#303030", fg="white").grid(row=1, column=0, padx=10, pady=5, sticky="e")
        self.num_deposits_b_entry = tk.Entry(input_frame, bg="#1d1d1e", fg="white")
        self.num_deposits_b_entry.grid(row=1, column=1, padx=10, pady=5, sticky="w")

        # Button to generate deposit fields or clear them
        self.generate_button = tk.Button(root, text="Generate Deposit Fields", command=self.toggle_fields)  # Default styling
        self.generate_button.pack(padx=10, pady=10)

        # Frame for Lists A and B with Scrollbars
        lists_frame = tk.Frame(root, bg="#303030")
        lists_frame.pack(padx=10, pady=10, fill="both", expand=True)

        # Scrollable Frame for List A
        self.scrollable_frame_a = ScrollableFrame(lists_frame)
        self.scrollable_frame_a.pack(side="left", fill="both", expand=True, padx=(0, 5))

        # Scrollable Frame for List B
        self.scrollable_frame_b = ScrollableFrame(lists_frame)
        self.scrollable_frame_b.pack(side="right", fill="both", expand=True, padx=(5, 0))

        # Labels for List A and List B
        label_a_header = tk.Label(self.scrollable_frame_a.scrollable_frame, text="List A Deposits:", font=("Arial", 14, "bold"), bg="#303030", fg="white")
        label_a_header.pack(anchor="w", pady=(0, 10))
        self.section_labels_a = [label_a_header]

        label_b_header = tk.Label(self.scrollable_frame_b.scrollable_frame, text="List B Deposits:", font=("Arial", 14, "bold"), bg="#303030", fg="white")
        label_b_header.pack(anchor="w", pady=(0, 10))
        self.section_labels_b = [label_b_header]

        # Placeholders for deposit entries
        self.list_a_entries = []
        self.list_b_entries = []
        self.verify_button = None
        self.results_label = None
        self.highlight_enabled = False
        self.selected_related_sets = []  # To store selected related sets with their indices
        self.deposit_to_pair_a = {}  # Mapping from List A deposit index to subset pair index
        self.deposit_to_pair_b = {}  # Mapping from List B deposit index to subset pair index

        # Internal tables for deposits
        self.table_a = {}  # e.g., {'A1': {'value': 3, 'status': 'Matched', 'related_set': 'R1'}}
        self.table_b = {}  # e.g., {'B1': {'value': 6, 'status': 'Matched', 'related_set': 'R1'}}

    def show_help(self):
        # Display a help message
        help_text = (
            "DepositsMatcher is a tool for accounting and bookkeeping to match "
            "deposit amounts between two lists (List A and List B). "
            "\n\nFeatures:\n"
            "- Input deposits for List A and List B.\n"
            "- Find matching subset totals between the two lists.\n"
            "- Highlight discrepancies if there are unmatched amounts.\n\n"
            "This tool helps users compare deposit entries to ensure accuracy.\n\n"
            "**Bulk Pasting:**\n"
            "After generating the deposit fields, you can paste multiple deposit values "
            "directly from a spreadsheet into List A or List B by clicking the 'Paste List A' or 'Paste List B' buttons."
        )
        messagebox.showinfo("Help - DepositsMatcher", help_text)

    def toggle_fields(self):
        # If fields are generated, clear them; otherwise, generate them
        if self.generate_button["text"] == "Generate Deposit Fields":
            self.generate_fields()
            self.generate_button.config(text="Clear Deposit Fields")
        else:
            self.clear_fields()
            self.generate_button.config(text="Generate Deposit Fields")

    def generate_fields(self):
        # Get the number of deposit fields to create for List A and List B
        try:
            num_deposits_a = int(self.num_deposits_a_entry.get())
            num_deposits_b = int(self.num_deposits_b_entry.get())
            if num_deposits_a <= 0 or num_deposits_b <= 0:
                raise ValueError("Number of deposits must be positive.")
        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid positive integers for the number of deposits.")
            return

        # Create deposit fields for List A
        for i in range(num_deposits_a):
            frame = tk.Frame(self.scrollable_frame_a.scrollable_frame, bg="#303030")
            frame.pack(anchor="w", pady=2, padx=5)

            label = tk.Label(frame, text=f"A{i+1}:", width=5, bg="#303030", fg="white", anchor="w")
            label.pack(side="left")

            entry = tk.Entry(frame, width=20, bg="#1d1d1e", fg="white")
            entry.pack(side="left", padx=(0, 10))
            # Bind events for highlighting
            entry.bind("<Enter>", lambda event, list_type="A": self.highlight_related(event, list_type))
            entry.bind("<Leave>", lambda event, list_type="A": self.clear_highlight())

            status_label = tk.Label(frame, text="", font=("Arial", 10, "bold"), bg="#303030", fg="white")
            status_label.pack(side="left")
            self.list_a_entries.append((entry, status_label))

        # Add 'Paste List A' Button if it does not already exist
        if not hasattr(self, "paste_a_button"):
            self.paste_a_button = tk.Button(self.scrollable_frame_a.scrollable_frame, text="Paste List A", command=self.paste_list_a)
            self.paste_a_button.pack(anchor="w", pady=(5, 10))

        # Create deposit fields for List B
        for i in range(num_deposits_b):
            frame = tk.Frame(self.scrollable_frame_b.scrollable_frame, bg="#303030")
            frame.pack(anchor="w", pady=2, padx=5)

            label = tk.Label(frame, text=f"B{i+1}:", width=5, bg="#303030", fg="white", anchor="w")
            label.pack(side="left")

            entry = tk.Entry(frame, width=20, bg="#1d1d1e", fg="white")
            entry.pack(side="left", padx=(0, 10))
            # Bind events for highlighting
            entry.bind("<Enter>", lambda event, list_type="B": self.highlight_related(event, list_type))
            entry.bind("<Leave>", lambda event, list_type="B": self.clear_highlight())

            status_label = tk.Label(frame, text="", font=("Arial", 10, "bold"), bg="#303030", fg="white")
            status_label.pack(side="left")
            self.list_b_entries.append((entry, status_label))

        # Add 'Paste List B' Button if it does not already exist
        if not hasattr(self, "paste_b_button"):
            self.paste_b_button = tk.Button(self.scrollable_frame_b.scrollable_frame, text="Paste List B", command=self.paste_list_b)
            self.paste_b_button.pack(anchor="w", pady=(5, 10))

        # Create the button to verify the sums
        if not self.verify_button:
            self.verify_button = tk.Button(self.root, text="Find Maximum Matching Sum", command=self.find_max_matching_sum)
            self.verify_button.pack(padx=10, pady=10)

        # Add the Export button only once
        if not hasattr(self, "export_button"):
            self.export_button = tk.Button(self.root, text="Export to Excel", command=self.export_to_excel)
            self.export_button.pack(padx=10, pady=10)


    def clear_fields(self):
        # Clear all dynamically created fields and reset state
        for entry, status_label in self.list_a_entries + self.list_b_entries:
            entry.master.destroy()  # Destroy the parent frame
        self.list_a_entries.clear()
        self.list_b_entries.clear()

        if self.verify_button:
            self.verify_button.destroy()
            self.verify_button = None

        if self.results_label:
            self.results_label.destroy()
            self.results_label = None

        # Reset tracking variables
        self.highlight_enabled = False
        self.selected_related_sets = []
        self.deposit_to_pair_a.clear()
        self.deposit_to_pair_b.clear()

        # Clear any text in the entry fields for deposit counts
        self.num_deposits_a_entry.delete(0, tk.END)
        self.num_deposits_b_entry.delete(0, tk.END)

        # Clear internal tables
        self.table_a.clear()
        self.table_b.clear()
        

    def paste_list_a(self):
        """
        Handle pasting multiple deposit values into List A.
        """
        try:
            clipboard_data = self.root.clipboard_get()
            # Split by newlines and possibly tabs
            lines = [line.strip() for line in clipboard_data.replace('\r', '').split('\n') if line.strip()]
            num_entries = len(self.list_a_entries)
            if len(lines) > num_entries:
                messagebox.showwarning("Paste Warning", f"Number of pasted values ({len(lines)}) exceeds the number of deposit entries ({num_entries}). Extra values will be ignored.")
            for i, entry_tuple in enumerate(self.list_a_entries):
                if i < len(lines):
                    entry, _ = entry_tuple
                    entry.delete(0, tk.END)
                    entry.insert(0, lines[i])
                else:
                    # Optionally, clear remaining entries or leave them as is
                    pass
        except tk.TclError:
            messagebox.showerror("Paste Error", "Clipboard does not contain valid text data.")

    def paste_list_b(self):
        """
        Handle pasting multiple deposit values into List B.
        """
        try:
            clipboard_data = self.root.clipboard_get()
            # Split by newlines and possibly tabs
            lines = [line.strip() for line in clipboard_data.replace('\r', '').split('\n') if line.strip()]
            num_entries = len(self.list_b_entries)
            if len(lines) > num_entries:
                messagebox.showwarning("Paste Warning", f"Number of pasted values ({len(lines)}) exceeds the number of deposit entries ({num_entries}). Extra values will be ignored.")
            for i, entry_tuple in enumerate(self.list_b_entries):
                if i < len(lines):
                    entry, _ = entry_tuple
                    entry.delete(0, tk.END)
                    entry.insert(0, lines[i])
                else:
                    # Optionally, clear remaining entries or leave them as is
                    pass
        except tk.TclError:
            messagebox.showerror("Paste Error", "Clipboard does not contain valid text data.")

    def find_max_matching_sum(self):
        # Check number of deposits and warn if more than 20
        num_deposits_a = len(self.list_a_entries)
        num_deposits_b = len(self.list_b_entries)
        if num_deposits_a > 20 or num_deposits_b > 20:
            if not messagebox.askokcancel("Warning", "You have more than 20 deposits. Calculations may take a while. Continue?"):
                return

        # Create and place the progress bar widget if not already created
        if not hasattr(self, 'progress_bar'):
            self.progress_bar = ttk.Progressbar(self.root, orient="horizontal", length=300, mode="determinate")
            self.progress_bar.pack(padx=10, pady=10)
        self.progress_bar["value"] = 0  # Reset progress bar
        self.progress_bar["maximum"] = 100  # Scale the progress bar from 0 to 100%
        self.root.update_idletasks()  # Ensure the UI updates to show the progress bar

        # Start the computation in a separate thread
        computation_thread = threading.Thread(target=self._run_matching_computation)
        computation_thread.start()

    def _run_matching_computation(self):
        try:
            # Perform the matching computation
            # Set up lists of deposit values
            deposits_a = [float(entry.get()) for entry, _ in self.list_a_entries]
            deposits_b = [float(entry.get()) for entry, _ in self.list_b_entries]

            # Populate internal tables
            self.table_a = {}
            for i, value in enumerate(deposits_a):
                self.table_a[f"A{i+1}"] = {'value': value, 'status': 'Unmatched', 'related_set': None}

            self.table_b = {}
            for i, value in enumerate(deposits_b):
                self.table_b[f"B{i+1}"] = {'value': value, 'status': 'Unmatched', 'related_set': None}

            # Generate all possible subsets for both lists with their deposit indices
            subsets_a = self.get_all_subsets(deposits_a, prefix='A')
            subsets_b = self.get_all_subsets(deposits_b, prefix='B')

            # Update progress bar as we find matching subset pairs
            self._update_progress(20)

            # Find all matching subset pairs (same sum)
            matching_subset_pairs = self.find_matching_subset_pairs(subsets_a, subsets_b)

            # Update progress bar
            self._update_progress(50)

            # Find the optimal combination of subset pairs
            optimal_matching = self.find_optimal_matching(matching_subset_pairs)

            # Update progress bar
            self._update_progress(80)

            # Assign related set identifiers and update internal tables
            related_set_id = 1
            for pair in optimal_matching:
                subset_a, subset_b, total = pair
                related_set = f"R{related_set_id}"
                related_set_id += 1

                # Update table for List A
                for deposit_id in subset_a['deposit_ids']:
                    self.table_a[deposit_id]['status'] = 'Matched'
                    self.table_a[deposit_id]['related_set'] = related_set

                # Update table for List B
                for deposit_id in subset_b['deposit_ids']:
                    self.table_b[deposit_id]['status'] = 'Matched'
                    self.table_b[deposit_id]['related_set'] = related_set

                # Store related sets for highlighting
                self.selected_related_sets.append({
                    'related_set': related_set,
                    'a_deposits': subset_a['deposit_ids'],
                    'b_deposits': subset_b['deposit_ids']
                })

            # Finalize progress bar
            self._update_progress(100)

            # Display results (UI updates must happen on the main thread)
            self.root.after(0, self._display_results)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred during computation: {e}")

    def _update_progress(self, value):
        """
        Update the progress bar and ensure the UI refreshes.
        """
        self.progress_bar["value"] = value
        self.root.update_idletasks()


    def _display_results(self):
        """
        Display the results on the UI after computation is complete.
        """
        total_matched_a = sum([self.table_a[deposit_id]['value'] for deposit_id in self.table_a if self.table_a[deposit_id]['status'] == 'Matched'])
        total_matched_b = sum([self.table_b[deposit_id]['value'] for deposit_id in self.table_b if self.table_b[deposit_id]['status'] == 'Matched'])

        result_message = f"Total Matched: {total_matched_a}\n"
        result_message += f"Total Unmatched - List A: {sum([self.table_a[deposit_id]['value'] for deposit_id in self.table_a if self.table_a[deposit_id]['status'] == 'Unmatched'])}\n"
        result_message += f"Total Unmatched - List B: {sum([self.table_b[deposit_id]['value'] for deposit_id in self.table_b if self.table_b[deposit_id]['status'] == 'Unmatched'])}\n\nMatched Subsets:\n"
        for idx, related_set in enumerate(self.selected_related_sets, 1):
            subset_a_ids = related_set['a_deposits']
            subset_b_ids = related_set['b_deposits']
            subset_a_values = [self.table_a[dep_id]['value'] for dep_id in subset_a_ids]
            subset_b_values = [self.table_b[dep_id]['value'] for dep_id in subset_b_ids]
            subset_a_str = ', '.join(map(str, subset_a_values))
            subset_b_str = ', '.join(map(str, subset_b_values))
            subset_sum = round(sum(subset_a_values), 2)
            result_message += f"Pair {idx}: List A [{subset_a_str}] <--> List B [{subset_b_str}] (Sum: {subset_sum})\n"

        if self.results_label:
            self.results_label.destroy()
        self.results_label = tk.Label(self.root, text=result_message, font=("Arial", 12, "bold"), justify=tk.LEFT, anchor="w", bg="#303030", fg="white")
        self.results_label.pack(padx=10, pady=10, fill="both", expand=True)

        # Update status labels for List A
        for i, (entry, status_label) in enumerate(self.list_a_entries):
            deposit_id = f"A{i+1}"
            status = self.table_a[deposit_id]['status']
            if status == "Matched":
                status_label.config(text="Matched", fg="green", bg="#303030")
            else:
                status_label.config(text="Unmatched", fg="red", bg="#303030")

        # Update status labels for List B
        for i, (entry, status_label) in enumerate(self.list_b_entries):
            deposit_id = f"B{i+1}"
            status = self.table_b[deposit_id]['status']
            if status == "Matched":
                status_label.config(text="Matched", fg="green", bg="#303030")
            else:
                status_label.config(text="Unmatched", fg="red", bg="#303030")

        # Enable highlighting
        self.highlight_enabled = True


    

    def get_all_subsets(self, deposits, prefix):
        """
        Generate all possible non-empty subsets for a list of deposits.
        Each subset includes the sum and the deposit IDs.
        """
        subsets = []
        for r in range(1, len(deposits) + 1):
            for subset in combinations(enumerate(deposits), r):
                deposit_indices = [f"{prefix}{i+1}" for i, _ in subset]
                subset_sum = round(sum([val for _, val in subset]), 10)  # Avoid floating-point issues
                subsets.append({
                    'deposit_ids': deposit_indices,
                    'sum': subset_sum
                })
        return subsets

    def find_matching_subset_pairs(self, subsets_a, subsets_b):
        """
        Find all subset pairs from A and B that have the same sum.
        """
        sum_to_subsets_a = {}
        for subset in subsets_a:
            sum_val = subset['sum']
            if sum_val not in sum_to_subsets_a:
                sum_to_subsets_a[sum_val] = []
            sum_to_subsets_a[sum_val].append(subset)

        sum_to_subsets_b = {}
        for subset in subsets_b:
            sum_val = subset['sum']
            if sum_val not in sum_to_subsets_b:
                sum_to_subsets_b[sum_val] = []
            sum_to_subsets_b[sum_val].append(subset)

        # Find matching sums
        matching_subset_pairs = []
        for sum_val in sum_to_subsets_a:
            if sum_val in sum_to_subsets_b:
                for subset_a in sum_to_subsets_a[sum_val]:
                    for subset_b in sum_to_subsets_b[sum_val]:
                        matching_subset_pairs.append((subset_a, subset_b, sum_val))
        return matching_subset_pairs

    def find_optimal_matching(self, matching_subset_pairs):
        """
        Find the optimal combination of subset pairs that maximizes the total sum
        without overlapping deposits. This version is iterative to avoid recursion limits.
        """
        best_matching = []
        best_total = 0

        # Convert matching_subset_pairs to include unique identifiers for each subset
        indexed_pairs = [(index, subset_a, subset_b, sum_val)
                         for index, (subset_a, subset_b, sum_val) in enumerate(matching_subset_pairs)]

        # Generate combinations iteratively to find non-overlapping matches
        for combination in product(*indexed_pairs):
            used_a = set()
            used_b = set()
            current_matching = []
            current_total = 0
            valid_combination = True

            for _, subset_a, subset_b, sum_val in combination:
                # Check if the subsets overlap with already used elements
                if any(dep_id in used_a for dep_id in subset_a['deposit_ids']) or \
                   any(dep_id in used_b for dep_id in subset_b['deposit_ids']):
                    valid_combination = False
                    break

                # Add to current matching if valid
                current_matching.append((subset_a, subset_b, sum_val))
                current_total += sum_val

                # Mark elements as used
                used_a.update(subset_a['deposit_ids'])
                used_b.update(subset_b['deposit_ids'])

            # Check if this combination has a higher total than the best found
            if valid_combination and current_total > best_total:
                best_total = current_total
                best_matching = current_matching

        return best_matching

    def paste_list_a(self):
        """
        Handle pasting multiple deposit values into List A.
        """
        try:
            clipboard_data = self.root.clipboard_get()
            # Split by newlines and possibly tabs
            lines = [line.strip() for line in clipboard_data.replace('\r', '').split('\n') if line.strip()]
            num_entries = len(self.list_a_entries)
            if len(lines) > num_entries:
                messagebox.showwarning("Paste Warning", f"Number of pasted values ({len(lines)}) exceeds the number of deposit entries ({num_entries}). Extra values will be ignored.")
            for i, entry_tuple in enumerate(self.list_a_entries):
                if i < len(lines):
                    entry, _ = entry_tuple
                    entry.delete(0, tk.END)
                    entry.insert(0, lines[i])
                else:
                    # Optionally, clear remaining entries or leave them as is
                    pass
        except tk.TclError:
            messagebox.showerror("Paste Error", "Clipboard does not contain valid text data.")

    def paste_list_b(self):
        """
        Handle pasting multiple deposit values into List B.
        """
        try:
            clipboard_data = self.root.clipboard_get()
            # Split by newlines and possibly tabs
            lines = [line.strip() for line in clipboard_data.replace('\r', '').split('\n') if line.strip()]
            num_entries = len(self.list_b_entries)
            if len(lines) > num_entries:
                messagebox.showwarning("Paste Warning", f"Number of pasted values ({len(lines)}) exceeds the number of deposit entries ({num_entries}). Extra values will be ignored.")
            for i, entry_tuple in enumerate(self.list_b_entries):
                if i < len(lines):
                    entry, _ = entry_tuple
                    entry.delete(0, tk.END)
                    entry.insert(0, lines[i])
                else:
                    # Optionally, clear remaining entries or leave them as is
                    pass
        except tk.TclError:
            messagebox.showerror("Paste Error", "Clipboard does not contain valid text data.")

    def find_max_matching_sum(self):
        # Find matching subsets and related sets logic
        try:
            deposits_a = [float(entry.get()) for entry, _ in self.list_a_entries]
            deposits_b = [float(entry.get()) for entry, _ in self.list_b_entries]
        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numbers for all deposits.")
            return

        # Populate internal tables
        self.table_a = {}
        for i, value in enumerate(deposits_a):
            self.table_a[f"A{i+1}"] = {'value': value, 'status': 'Unmatched', 'related_set': None}

        self.table_b = {}
        for i, value in enumerate(deposits_b):
            self.table_b[f"B{i+1}"] = {'value': value, 'status': 'Unmatched', 'related_set': None}

        # Generate all possible subsets for both lists with their deposit indices
        subsets_a = self.get_all_subsets(deposits_a, prefix='A')
        subsets_b = self.get_all_subsets(deposits_b, prefix='B')

        # Find all matching subset pairs (same sum)
        matching_subset_pairs = self.find_matching_subset_pairs(subsets_a, subsets_b)

        # Find the optimal combination of subset pairs
        optimal_matching = self.find_optimal_matching(matching_subset_pairs)

        # Assign related set identifiers and update internal tables
        related_set_id = 1
        for pair in optimal_matching:
            subset_a, subset_b, total = pair
            related_set = f"R{related_set_id}"
            related_set_id += 1

            # Update table for List A
            for deposit_id in subset_a['deposit_ids']:
                self.table_a[deposit_id]['status'] = 'Matched'
                self.table_a[deposit_id]['related_set'] = related_set

            # Update table for List B
            for deposit_id in subset_b['deposit_ids']:
                self.table_b[deposit_id]['status'] = 'Matched'
                self.table_b[deposit_id]['related_set'] = related_set

            # Store related sets for highlighting
            self.selected_related_sets.append({
                'related_set': related_set,
                'a_deposits': subset_a['deposit_ids'],
                'b_deposits': subset_b['deposit_ids']
            })

        # Calculate total matched sums
        total_matched_a = sum([self.table_a[deposit_id]['value'] for deposit_id in self.table_a if self.table_a[deposit_id]['status'] == 'Matched'])
        total_matched_b = sum([self.table_b[deposit_id]['value'] for deposit_id in self.table_b if self.table_b[deposit_id]['status'] == 'Matched'])

        # Prepare result message
        result_message = f"Total Matched: {total_matched_a}\n"
        result_message += f"Total Unmatched - List A: {sum([self.table_a[deposit_id]['value'] for deposit_id in self.table_a if self.table_a[deposit_id]['status'] == 'Unmatched'])}\n"
        result_message += f"Total Unmatched - List B: {sum([self.table_b[deposit_id]['value'] for deposit_id in self.table_b if self.table_b[deposit_id]['status'] == 'Unmatched'])}\n\nMatched Subsets:\n"
        for idx, related_set in enumerate(self.selected_related_sets, 1):
            subset_a_ids = related_set['a_deposits']
            subset_b_ids = related_set['b_deposits']
            subset_a_values = [self.table_a[dep_id]['value'] for dep_id in subset_a_ids]
            subset_b_values = [self.table_b[dep_id]['value'] for dep_id in subset_b_ids]
            subset_a_str = ', '.join(map(str, subset_a_values))
            subset_b_str = ', '.join(map(str, subset_b_values))
            subset_sum = round(sum(subset_a_values), 2)  # Assuming subset_a_sum == subset_b_sum
            result_message += f"Pair {idx}: List A [{subset_a_str}] <--> List B [{subset_b_str}] (Sum: {subset_sum})\n"

        if self.results_label:
            self.results_label.destroy()
        self.results_label = tk.Label(self.root, text=result_message, font=("Arial", 12, "bold"), justify=tk.LEFT, anchor="w", bg="#303030", fg="white")
        self.results_label.pack(padx=10, pady=10, fill="both", expand=True)

        # Update status labels for List A
        for i, (entry, status_label) in enumerate(self.list_a_entries):
            deposit_id = f"A{i+1}"
            status = self.table_a[deposit_id]['status']
            if status == "Matched":
                status_label.config(text="Matched", fg="green", bg="#303030")
            else:
                status_label.config(text="Unmatched", fg="red", bg="#303030")

        # Update status labels for List B
        for i, (entry, status_label) in enumerate(self.list_b_entries):
            deposit_id = f"B{i+1}"
            status = self.table_b[deposit_id]['status']
            if status == "Matched":
                status_label.config(text="Matched", fg="green", bg="#303030")
            else:
                status_label.config(text="Unmatched", fg="red", bg="#303030")

        # Enable highlighting
        self.highlight_enabled = True

    def get_all_subsets(self, deposits, prefix):
        """
        Generate all possible non-empty subsets for a list of deposits.
        Each subset includes the sum and the deposit IDs.
        """
        subsets = []
        for r in range(1, len(deposits) + 1):
            for subset in combinations(enumerate(deposits), r):
                deposit_indices = [f"{prefix}{i+1}" for i, _ in subset]
                subset_sum = round(sum([val for _, val in subset]), 10)  # Avoid floating-point issues
                subsets.append({
                    'deposit_ids': deposit_indices,
                    'sum': subset_sum
                })
        return subsets

    def find_matching_subset_pairs(self, subsets_a, subsets_b):
        """
        Find all subset pairs from A and B that have the same sum.
        """
        sum_to_subsets_a = {}
        for subset in subsets_a:
            sum_val = subset['sum']
            if sum_val not in sum_to_subsets_a:
                sum_to_subsets_a[sum_val] = []
            sum_to_subsets_a[sum_val].append(subset)

        sum_to_subsets_b = {}
        for subset in subsets_b:
            sum_val = subset['sum']
            if sum_val not in sum_to_subsets_b:
                sum_to_subsets_b[sum_val] = []
            sum_to_subsets_b[sum_val].append(subset)

        # Find matching sums
        matching_subset_pairs = []
        for sum_val in sum_to_subsets_a:
            if sum_val in sum_to_subsets_b:
                for subset_a in sum_to_subsets_a[sum_val]:
                    for subset_b in sum_to_subsets_b[sum_val]:
                        matching_subset_pairs.append((subset_a, subset_b, sum_val))
        return matching_subset_pairs

    def find_optimal_matching(self, matching_subset_pairs):
        """
        Find the optimal combination of subset pairs that maximizes the total sum
        without overlapping deposits. This version is iterative to avoid recursion limits.
        """
        best_matching = []
        best_total = 0

        # We will check each combination of matching subset pairs
        # Since combinations can be large, use each pair individually in a loop to avoid deep recursion
        for i in range(len(matching_subset_pairs)):
            # Track used deposits to avoid overlaps
            used_a = set()
            used_b = set()
            current_matching = []
            current_total = 0

            # Go through each subset pair starting from the i-th element
            for subset_a, subset_b, sum_val in matching_subset_pairs[i:]:
                # Check if this subset pair has any overlap with already used elements
                if any(dep_id in used_a for dep_id in subset_a['deposit_ids']) or \
                   any(dep_id in used_b for dep_id in subset_b['deposit_ids']):
                    continue  # Skip this pair if there's an overlap

                # Add this subset pair to the current matching
                current_matching.append((subset_a, subset_b, sum_val))
                current_total += sum_val

                # Mark elements as used
                used_a.update(subset_a['deposit_ids'])
                used_b.update(subset_b['deposit_ids'])

            # Check if this combination has a higher total than the best found
            if current_total > best_total:
                best_total = current_total
                best_matching = current_matching

        return best_matching


    def highlight_related(self, event, list_type):
        if not self.highlight_enabled:
            return

        widget = event.widget
        try:
            if list_type == "A":
                # Find the deposit ID of the hovered entry in List A
                deposit_id = self.get_deposit_id(self.list_a_entries, widget, 'A')
                if not deposit_id:
                    return
                related_set = self.table_a[deposit_id]['related_set']
                if not related_set:
                    return  # Unmatched; no highlighting needed
                # Retrieve the related set details
                related_set_details = next((rs for rs in self.selected_related_sets if rs['related_set'] == related_set), None)
                if not related_set_details:
                    return
                # Highlight the related deposits
                self._highlight_subset(related_set_details['a_deposits'], related_set_details['b_deposits'])
            else:
                # Find the deposit ID of the hovered entry in List B
                deposit_id = self.get_deposit_id(self.list_b_entries, widget, 'B')
                if not deposit_id:
                    return
                related_set = self.table_b[deposit_id]['related_set']
                if not related_set:
                    return  # Unmatched; no highlighting needed
                # Retrieve the related set details
                related_set_details = next((rs for rs in self.selected_related_sets if rs['related_set'] == related_set), None)
                if not related_set_details:
                    return
                # Highlight the related deposits
                self._highlight_subset(related_set_details['a_deposits'], related_set_details['b_deposits'])
        except Exception as e:
            print("Error in highlight_related:", e)

    def get_deposit_id(self, entries_list, widget, prefix):
        """
        Helper function to get the deposit ID (e.g., 'A1', 'B2') based on the widget.
        """
        for i, (entry, _) in enumerate(entries_list):
            if entry == widget:
                return f"{prefix}{i+1}"
        return None

    def _highlight_subset(self, indices_a, indices_b):
        # First, reset all entries to their default colors
        self.clear_highlight()

        # Define a highlight color
        highlight_color = "#e7786b"  # Reddish-orange

        # Highlight List A entries
        for dep_id in indices_a:
            index = self.get_index_from_deposit_id(dep_id, 'A')
            if index is not None:
                self.list_a_entries[index][0].config(bg=highlight_color)

        # Highlight List B entries
        for dep_id in indices_b:
            index = self.get_index_from_deposit_id(dep_id, 'B')
            if index is not None:
                self.list_b_entries[index][0].config(bg=highlight_color)

    def get_index_from_deposit_id(self, deposit_id, list_type):
        """
        Helper function to get the index of the deposit in the entries list based on deposit ID.
        """
        if list_type == 'A':
            return int(deposit_id[1:]) - 1  # Convert 'A1' to 0
        elif list_type == 'B':
            return int(deposit_id[1:]) - 1  # Convert 'B1' to 0
        return None

    def clear_highlight(self):
        if not self.highlight_enabled:
            return

        # Reset background colors for List A entries
        for i, (entry, status_label) in enumerate(self.list_a_entries):
            deposit_id = f"A{i+1}"
            entry.config(bg="#1d1d1e")  # Reset to default background

        # Reset background colors for List B entries
        for i, (entry, status_label) in enumerate(self.list_b_entries):
            deposit_id = f"B{i+1}"
            entry.config(bg="#1d1d1e")  # Reset to default background
            
    def export_to_excel(self):
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Deposits Matching"

        # Set headers
        ws['A1'] = "List A"
        ws['B1'] = "List B"

        # Define color fonts using hexadecimal values for red and green
        matched_font = Font(color="00FF00")  # Green text
        unmatched_font = Font(color="FF0000")  # Red text

        # Populate columns with data from List A and List B with conditional formatting
        for i, (entry, status_label) in enumerate(self.list_a_entries, start=2):
            deposit_id = f"A{i-1}"
            value = self.table_a[deposit_id]['value']
            status = self.table_a[deposit_id]['status']
            font = matched_font if status == "Matched" else unmatched_font
            cell = ws.cell(row=i, column=1, value=value)
            cell.font = font

        for i, (entry, status_label) in enumerate(self.list_b_entries, start=2):
            deposit_id = f"B{i-1}"
            value = self.table_b[deposit_id]['value']
            status = self.table_b[deposit_id]['status']
            font = matched_font if status == "Matched" else unmatched_font
            cell = ws.cell(row=i, column=2, value=value)
            cell.font = font

        # Open a dialog for saving the file
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb.save(file_path)
            messagebox.showinfo("Export Successful", f"Spreadsheet saved as {file_path}")


# Main loop
if __name__ == "__main__":
    root = tk.Tk()
    app = DepositsMatcherApp(root)
    root.mainloop()
